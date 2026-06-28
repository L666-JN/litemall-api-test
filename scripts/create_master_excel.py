"""
一次性脚本：创建 test_design/litemall_testcases.xlsx 统一用例文件
包含 auth 5 个 sheet（已有数据）+ goods 5 个 sheet（带表头，待填写）

运行: python scripts/create_master_excel.py
"""
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================
#  样式
# ============================================================
HEADER_FONT = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

CASE_FONT = Font(name="微软雅黑", size=10)
CASE_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

P0_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
P1_FILL = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
P2_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

HEADERS = [
    "用例编号", "用例ID", "模块", "用例描述", "优先级",
    "接口路径", "请求方法", "需要认证", "请求参数", "请求体",
    "预期HTTP状态码", "预期errno", "预期errno不为",
    "预期errmsg", "预期errmsg包含", "预期data有值", "预期data字段",
]

COL_WIDTHS = [18, 28, 12, 42, 8, 30, 10, 12, 30, 55, 16, 12, 14, 22, 22, 16, 38]

COUNTER_MAP = {
    "登录":     ("TC-AUTH-LOGIN", "auth"),
    "登出":     ("TC-AUTH-LGOUT", "auth"),
    "用户信息": ("TC-AUTH-INFO",  "auth"),
    "注册":     ("TC-AUTH-REG",   "auth"),
    "密码重置": ("TC-AUTH-RSET",  "auth"),
    "商品列表": ("TC-GOODS-LIST", "goods"),
    "商品详情": ("TC-GOODS-DTL",  "goods"),
    "商品分类": ("TC-GOODS-CAT",  "goods"),
    "相关商品": ("TC-GOODS-REL",  "goods"),
    "商品统计": ("TC-GOODS-CNT",  "goods"),
    "购物车列表": ("TC-CART-LIST",  "cart"),
    "购物车添加": ("TC-CART-ADD",  "cart"),
    "快速添加": ("TC-CART-FADD",  "cart"),
    "购物车更新": ("TC-CART-UPD",  "cart"),
    "购物车选中": ("TC-CART-CHK",  "cart"),
    "购物车删除": ("TC-CART-DEL",  "cart"),
    "购物车数量": ("TC-CART-CNT",  "cart"),
    "购物车结算": ("TC-CART-CHOT", "cart"),
    "地址列表": ("TC-ADDR-LIST",  "address"),
    "地址详情": ("TC-ADDR-DTL",  "address"),
    "地址保存": ("TC-ADDR-SAVE", "address"),
    "地址删除": ("TC-ADDR-DEL",  "address"),
    "订单列表": ("TC-ORDR-LIST",  "order"),
    "订单详情": ("TC-ORDR-DTL",  "order"),
    "订单提交": ("TC-ORDR-SUB",  "order"),
    "订单取消": ("TC-ORDR-CANC", "order"),
    "确认收货": ("TC-ORDR-CONF", "order"),
    "订单删除": ("TC-ORDR-DEL",  "order"),
    "订单评价": ("TC-ORDR-COMM", "order"),
    "订单商品": ("TC-ORDR-GOOD", "order"),
    "首页": ("TC-HOME", "home"),
}


# ============================================================
#  用例数据
# ============================================================

ALL_CASES = [
    # ── 登录 ──
    ["登录", "login_success", "正确的用户名和密码登录", "P0",
     "/wx/auth/login", "POST", "否", "",
     '{"username": "user123", "password": "user123"}',
     200, 0, "", "成功", "", "是", "token, userInfo, userInfo.nickName"],

    ["登录", "login_wrong_password", "密码错误登录", "P0",
     "/wx/auth/login", "POST", "否", "",
     '{"username": "user123", "password": "wrong_password"}',
     200, 700, "", "", "", "否", ""],

    ["登录", "login_nonexistent_user", "不存在的用户名登录", "P1",
     "/wx/auth/login", "POST", "否", "",
     '{"username": "no_such_user_999", "password": "123456"}',
     200, 700, "", "", "账号不存在", "否", ""],

    ["登录", "login_empty_username", "用户名为空登录", "P0",
     "/wx/auth/login", "POST", "否", "",
     '{"username": "", "password": "user123"}',
     200, "", "0", "", "", "否", ""],

    ["登录", "login_empty_password", "密码为空登录", "P0",
     "/wx/auth/login", "POST", "否", "",
     '{"username": "user123", "password": ""}',
     200, "", "0", "", "", "否", ""],

    ["登录", "login_empty_both", "用户名和密码都为空登录", "P1",
     "/wx/auth/login", "POST", "否", "",
     '{"username": "", "password": ""}',
     200, "", "0", "", "", "否", ""],

    ["登录", "login_username_with_spaces", "用户名前后带空格登录（边界）", "P2",
     "/wx/auth/login", "POST", "否", "",
     '{"username": "  user123  ", "password": "user123"}',
     200, "", "", "", "", "否", ""],

    ["登录", "login_sql_injection", "SQL注入尝试登录（安全测试）", "P1",
     "/wx/auth/login", "POST", "否", "",
     '{"username": "admin\' OR \'1\'=\'1", "password": "admin\' OR \'1\'=\'1"}',
     200, "", "0", "", "", "否", ""],

    ["登录", "login_xss_attempt", "XSS脚本注入尝试登录（安全测试）", "P2",
     "/wx/auth/login", "POST", "否", "",
     '{"username": "<script>alert(1)</script>", "password": "<script>alert(1)</script>"}',
     200, "", "0", "", "", "否", ""],

    # ── 登出 ──
    ["登出", "logout_with_token", "带有效token登出", "P0",
     "/wx/auth/logout", "POST", "是", "", "",
     200, 0, "", "成功", "", "否", ""],

    ["登出", "logout_no_token", "不带token登出（应拒绝）", "P0",
     "/wx/auth/logout", "POST", "否", "", "",
     200, 501, "", "请登录", "", "否", ""],

    ["登出", "logout_invalid_token", "带伪造的无效token登出（应拒绝）", "P1",
     "/wx/auth/logout", "POST", "否（手动设置假token）", "", "",
     200, 501, "", "请登录", "", "否", ""],

    # ── 用户信息 ──
    ["用户信息", "user_info_with_token", "带有效token获取用户信息", "P0",
     "/wx/auth/info", "GET", "是", "", "",
     200, 0, "", "成功", "", "是", "nickName, avatar, gender, mobile"],

    ["用户信息", "user_info_no_token", "不带token获取用户信息（应拒绝）", "P0",
     "/wx/auth/info", "GET", "否", "", "",
     200, 501, "", "请登录", "", "否", ""],

    ["用户信息", "user_info_invalid_token", "带伪造的无效token获取用户信息（应拒绝）", "P1",
     "/wx/auth/info", "GET", "否（手动设置假token）", "", "",
     200, 501, "", "", "", "否", ""],

    # ── 注册 ──
    ["注册", "register_missing_wxcode", "缺少wxCode字段注册（wxCode需通过regCaptcha获取）", "P1",
     "/wx/auth/register", "POST", "否", "",
     '{"username": "testuser_reg", "password": "Test1234!", "mobile": "13800138001"}',
     200, "", "0", "", "", "否", ""],

    ["注册", "register_empty_username", "空用户名注册（应拒绝）", "P1",
     "/wx/auth/register", "POST", "否", "",
     '{"username": "", "password": "Test1234!", "mobile": "13800138002"}',
     200, "", "0", "", "", "否", ""],

    ["注册", "register_empty_password", "空密码注册（应拒绝）", "P1",
     "/wx/auth/register", "POST", "否", "",
     '{"username": "testuser_nopwd", "password": "", "mobile": "13800138003"}',
     200, "", "0", "", "", "否", ""],

    ["注册", "register_empty_mobile", "空手机号注册（应拒绝）", "P1",
     "/wx/auth/register", "POST", "否", "",
     '{"username": "testuser_nomob", "password": "Test1234!", "mobile": ""}',
     200, "", "0", "", "", "否", ""],

    ["注册", "register_duplicate_username", "用已存在的用户名注册（应拒绝）", "P0",
     "/wx/auth/register", "POST", "否", "",
     '{"username": "user123", "password": "Test1234!", "mobile": "13800138004"}',
     200, "", "0", "", "", "否", ""],

    ["注册", "register_invalid_mobile", "格式非法的手机号注册（应拒绝）", "P1",
     "/wx/auth/register", "POST", "否", "",
     '{"username": "testuser_badmob", "password": "Test1234!", "mobile": "12345"}',
     200, "", "0", "", "", "否", ""],

    ["注册", "register_short_password", "密码过短注册（应拒绝，如后端有校验）", "P2",
     "/wx/auth/register", "POST", "否", "",
     '{"username": "testuser_shortpwd", "password": "12", "mobile": "13800138005"}',
     200, "", "0", "", "", "否", ""],

    # ── 密码重置 ──
    ["密码重置", "reset_with_valid_data", "有效数据重置密码", "P2",
     "/wx/auth/reset", "POST", "否", "",
     '{"password": "NewPass123!", "mobile": "13800138001", "code": "1234"}',
     200, "", "", "", "", "否", ""],

    ["密码重置", "reset_empty_password", "空密码重置（应拒绝）", "P2",
     "/wx/auth/reset", "POST", "否", "",
     '{"password": "", "mobile": "13800138001", "code": "1234"}',
     200, "", "0", "", "", "否", ""],

    ["密码重置", "reset_empty_mobile", "空手机号重置（应拒绝）", "P2",
     "/wx/auth/reset", "POST", "否", "",
     '{"password": "NewPass123!", "mobile": "", "code": "1234"}',
     200, "", "0", "", "", "否", ""],

    # ── 商品列表 ──
    ["商品列表", "goods_list_default", "默认商品列表（验证分页字段）", "P0",
     "/wx/goods/list", "GET", "否", "", "",
     200, 0, "", "成功", "", "是", "total, pages, limit, page, list"],

    ["商品列表", "goods_list_page2_size3", "第2页每页3条", "P0",
     "/wx/goods/list", "GET", "否", "page=2&size=3", "",
     200, 0, "", "", "", "是", "list"],

    ["商品列表", "goods_list_keyword_valid", "搜索关键词\"茶\"", "P1",
     "/wx/goods/list", "GET", "否", "keyword=茶", "",
     200, 0, "", "", "", "是", "list, filterCategoryList"],

    ["商品列表", "goods_list_keyword_none", "搜索无结果关键词", "P1",
     "/wx/goods/list", "GET", "否", "keyword=xyznonexist999", "",
     200, 0, "", "", "", "是", "list"],

    ["商品列表", "goods_list_is_new", "筛选新品", "P1",
     "/wx/goods/list", "GET", "否", "isNew=true", "",
     200, 0, "", "", "", "是", "list"],

    ["商品列表", "goods_list_is_hot", "筛选热门", "P1",
     "/wx/goods/list", "GET", "否", "isHot=true", "",
     200, 0, "", "", "", "是", "list"],

    ["商品列表", "goods_list_empty_category", "空分类（无商品）", "P1",
     "/wx/goods/list", "GET", "否", "categoryId=0", "",
     200, 0, "", "", "", "是", "list"],

    ["商品列表", "goods_list_sort_invalid", "无效排序字段", "P2",
     "/wx/goods/list", "GET", "否", "sort=invalid_field", "",
     200, 402, "", "", "", "否", ""],

    # ── 商品详情 ──
    ["商品详情", "goods_detail_missing_id", "不传id", "P0",
     "/wx/goods/detail", "GET", "否", "", "",
     200, 402, "", "id不能为null", "", "否", ""],

    ["商品详情", "goods_detail_zero_id", "id=0", "P1",
     "/wx/goods/detail", "GET", "否", "id=0", "",
     200, 502, "", "", "", "否", ""],

    ["商品详情", "goods_detail_nonexistent", "id=99999不存在", "P1",
     "/wx/goods/detail", "GET", "否", "id=99999", "",
     200, 502, "", "", "", "否", ""],

    # ── 商品分类 ──
    ["商品分类", "goods_category_missing_id", "不传id", "P0",
     "/wx/goods/category", "GET", "否", "", "",
     200, 402, "", "id不能为null", "", "否", ""],

    ["商品分类", "goods_category_nonexistent", "id=99999不存在", "P1",
     "/wx/goods/category", "GET", "否", "id=99999", "",
     200, 502, "", "", "", "否", ""],

    # ── 相关商品 ──
    ["相关商品", "goods_related_missing_id", "不传id", "P0",
     "/wx/goods/related", "GET", "否", "", "",
     200, 402, "", "id不能为null", "", "否", ""],

    ["相关商品", "goods_related_invalid_id", "id=99999不存在", "P1",
     "/wx/goods/related", "GET", "否", "id=99999", "",
     200, 402, "", "", "", "否", ""],

    # ── 商品统计 ──
    ["商品统计", "goods_count_default", "获取商品总数", "P0",
     "/wx/goods/count", "GET", "否", "", "",
     200, 0, "", "成功", "", "是", ""],

    # ── 购物车列表 ──
    ["购物车列表", "cart_index_success", "认证用户查看购物车列表", "P0",
     "/wx/cart/index", "GET", "是", "", "",
     200, 0, "", "成功", "", "是", "cartTotal, cartList"],

    ["购物车列表", "cart_index_unauthorized", "未认证查看购物车列表", "P0",
     "/wx/cart/index", "GET", "否", "", "",
     200, 501, "", "", "", "否", ""],

    # ── 购物车添加 ──
    ["购物车添加", "cart_add_unauthorized", "未认证添加商品到购物车", "P0",
     "/wx/cart/add", "POST", "否", "",
     '{"goodsId": 1006002, "number": 1, "productId": 7}',
     200, 501, "", "", "", "否", ""],

    ["购物车添加", "cart_add_missing_goodsId", "缺少goodsId添加购物车", "P1",
     "/wx/cart/add", "POST", "是", "",
     '{"number": 1, "productId": 7}',
     200, "", "0", "", "", "否", ""],

    ["购物车添加", "cart_add_missing_productId", "缺少productId添加购物车", "P1",
     "/wx/cart/add", "POST", "是", "",
     '{"goodsId": 1006002, "number": 1}',
     200, "", "0", "", "", "否", ""],

    ["购物车添加", "cart_add_missing_number", "缺少number添加购物车", "P1",
     "/wx/cart/add", "POST", "是", "",
     '{"goodsId": 1006002, "productId": 7}',
     200, "", "0", "", "", "否", ""],

    ["购物车添加", "cart_add_zero_number", "number=0添加购物车", "P1",
     "/wx/cart/add", "POST", "是", "",
     '{"goodsId": 1006002, "number": 0, "productId": 7}',
     200, "", "0", "", "", "否", ""],

    ["购物车添加", "cart_add_nonexistent_goods", "不存在的商品ID添加购物车", "P1",
     "/wx/cart/add", "POST", "是", "",
     '{"goodsId": 999999, "number": 1, "productId": 7}',
     200, "", "0", "", "", "否", ""],

    ["购物车添加", "cart_add_nonexistent_product", "存在商品但产品规格不存在", "P1",
     "/wx/cart/add", "POST", "是", "",
     '{"goodsId": 1006002, "number": 1, "productId": 999999}',
     200, "", "0", "", "", "否", ""],

    # ── 快速添加 ──
    ["快速添加", "cart_fastadd_unauthorized", "未认证快速添加商品到购物车", "P0",
     "/wx/cart/fastadd", "POST", "否", "",
     '{"goodsId": 1006002, "number": 1, "productId": 7}',
     200, 501, "", "", "", "否", ""],

    ["快速添加", "cart_fastadd_missing_goodsId", "缺少goodsId快速添加", "P1",
     "/wx/cart/fastadd", "POST", "是", "",
     '{"number": 1, "productId": 7}',
     200, "", "0", "", "", "否", ""],

    ["快速添加", "cart_fastadd_nonexistent_goods", "不存在的商品ID快速添加", "P1",
     "/wx/cart/fastadd", "POST", "是", "",
     '{"goodsId": 999999, "number": 1, "productId": 7}',
     200, "", "0", "", "", "否", ""],

    # ── 购物车更新 ──
    ["购物车更新", "cart_update_unauthorized", "未认证更新购物车商品数量", "P0",
     "/wx/cart/update", "POST", "否", "",
     '{"id": 1, "number": 3, "goodsId": 1006002, "productId": 7}',
     200, 501, "", "", "", "否", ""],

    ["购物车更新", "cart_update_missing_goodsId", "缺少goodsId更新购物车", "P1",
     "/wx/cart/update", "POST", "是", "",
     '{"id": 1, "number": 3, "productId": 7}',
     200, "", "0", "", "", "否", ""],

    ["购物车更新", "cart_update_nonexistent", "更新不存在的购物车条目", "P1",
     "/wx/cart/update", "POST", "是", "",
     '{"id": 99999, "number": 3, "goodsId": 1006002, "productId": 7}',
     200, "", "0", "", "", "否", ""],

    # ── 购物车选中 ──
    ["购物车选中", "cart_checked_unauthorized", "未认证变更购物车选中状态", "P0",
     "/wx/cart/checked", "POST", "否", "",
     '{"isChecked": 1, "productIds": [7]}',
     200, 501, "", "", "", "否", ""],

    ["购物车选中", "cart_checked_nonexistent_product", "选中不存在的productId（无操作返回成功）", "P1",
     "/wx/cart/checked", "POST", "是", "",
     '{"isChecked": 1, "productIds": [999999]}',
     200, 0, "", "成功", "", "是", ""],

    # ── 购物车删除 ──
    ["购物车删除", "cart_delete_unauthorized", "未认证删除购物车商品", "P0",
     "/wx/cart/delete", "POST", "否", "",
     '{"productIds": [7]}',
     200, 501, "", "", "", "否", ""],

    ["购物车删除", "cart_delete_nonexistent_product", "删除不存在的productId（无操作返回成功）", "P1",
     "/wx/cart/delete", "POST", "是", "",
     '{"productIds": [999999]}',
     200, 0, "", "成功", "", "是", ""],

    # ── 购物车数量 ──
    ["购物车数量", "cart_count_no_auth", "未认证查询购物车数量（公开接口返回0）", "P0",
     "/wx/cart/goodscount", "GET", "否", "", "",
     200, 0, "", "成功", "", "是", ""],

    ["购物车数量", "cart_count_success", "认证用户查询购物车商品数量", "P0",
     "/wx/cart/goodscount", "GET", "是", "", "",
     200, 0, "", "成功", "", "是", ""],

    # ── 购物车结算 ──
    ["购物车结算", "cart_checkout_unauthorized", "未认证购物车结算", "P0",
     "/wx/cart/checkout", "GET", "否",
     "cartId=0&addressId=0&couponId=0&userCouponId=0&grouponRulesId=0", "",
     200, 501, "", "", "", "否", ""],

    ["购物车结算", "cart_checkout_default_params", "默认参数购物车结算（空购物车也有结算数据）", "P1",
     "/wx/cart/checkout", "GET", "是", "", "",
     200, 0, "", "成功", "", "是", ""],

    # ── 地址列表 ──
    ["地址列表", "address_list_success", "认证用户查看地址列表", "P0",
     "/wx/address/list", "GET", "是", "", "",
     200, 0, "", "成功", "", "是", "total, pages, list"],

    ["地址列表", "address_list_unauthorized", "未认证查看地址列表", "P0",
     "/wx/address/list", "GET", "否", "", "",
     200, 501, "", "", "", "否", ""],

    # ── 地址详情 ──
    ["地址详情", "address_detail_unauthorized", "未认证查看地址详情", "P0",
     "/wx/address/detail", "GET", "否", "id=1", "",
     200, 501, "", "", "", "否", ""],

    ["地址详情", "address_detail_missing_id", "不传id查地址详情", "P0",
     "/wx/address/detail", "GET", "是", "", "",
     200, 402, "", "id不能为null", "", "否", ""],

    ["地址详情", "address_detail_nonexistent", "不存在的地址id", "P1",
     "/wx/address/detail", "GET", "是", "id=99999", "",
     200, 402, "", "", "", "否", ""],

    # ── 地址保存 ──
    ["地址保存", "address_save_unauthorized", "未认证保存地址", "P0",
     "/wx/address/save", "POST", "否", "",
     '{"name": "Test", "tel": "13800138000", "province": "北京市", "city": "北京市", "county": "东城区", "addressDetail": "测试路1号", "areaCode": "110101", "isDefault": false}',
     200, 501, "", "", "", "否", ""],

    ["地址保存", "address_save_empty_body", "空请求体保存地址", "P1",
     "/wx/address/save", "POST", "是", "", "{}",
     200, 401, "", "", "", "否", ""],

    ["地址保存", "address_save_missing_fields", "仅传name缺少必填字段", "P1",
     "/wx/address/save", "POST", "是", "",
     '{"name": "test"}',
     200, 401, "", "", "", "否", ""],

    ["地址保存", "address_save_update_nonexistent", "更新不存在的地址id", "P1",
     "/wx/address/save", "POST", "是", "",
     '{"id": 99999, "name": "Ghost", "tel": "13800138000", "province": "北京市", "city": "北京市", "county": "东城区", "addressDetail": "不存在", "areaCode": "110101"}',
     200, 401, "", "", "", "否", ""],

    # ── 地址删除 ──
    ["地址删除", "address_delete_unauthorized", "未认证删除地址", "P0",
     "/wx/address/delete", "POST", "否", "",
     '{"id": 1}',
     200, 501, "", "", "", "否", ""],

    ["地址删除", "address_delete_empty_body", "空请求体删除地址", "P1",
     "/wx/address/delete", "POST", "是", "", "{}",
     200, 401, "", "", "", "否", ""],

    ["地址删除", "address_delete_nonexistent", "删除不存在的地址", "P1",
     "/wx/address/delete", "POST", "是", "",
     '{"id": 99999}',
     200, 402, "", "", "", "否", ""],

    # ── 订单列表 ──
    ["订单列表", "order_list_success", "认证用户查看订单列表", "P0",
     "/wx/order/list", "GET", "是", "", "",
     200, 0, "", "成功", "", "是", "total, pages, list"],

    ["订单列表", "order_list_unauthorized", "未认证查看订单列表", "P0",
     "/wx/order/list", "GET", "否", "", "",
     200, 501, "", "", "", "否", ""],

    # ── 订单详情 ──
    ["订单详情", "order_detail_unauthorized", "未认证查看订单详情", "P0",
     "/wx/order/detail", "GET", "否", "orderId=1", "",
     200, 501, "", "", "", "否", ""],

    ["订单详情", "order_detail_missing_id", "不传orderId查订单详情", "P0",
     "/wx/order/detail", "GET", "是", "", "",
     200, 402, "", "orderId不能为null", "", "否", ""],

    ["订单详情", "order_detail_nonexistent", "不存在的订单id", "P1",
     "/wx/order/detail", "GET", "是", "orderId=99999", "",
     200, 720, "", "", "", "否", ""],

    # ── 订单提交 ──
    ["订单提交", "order_submit_unauthorized", "未认证提交订单", "P0",
     "/wx/order/submit", "POST", "否", "",
     '{"cartId": 0, "addressId": 1, "couponId": -1, "userCouponId": -1, "grouponRulesId": 0, "message": "test"}',
     200, 501, "", "", "", "否", ""],

    ["订单提交", "order_submit_empty_body", "空请求体提交订单", "P1",
     "/wx/order/submit", "POST", "是", "", "{}",
     200, 401, "", "", "", "否", ""],

    # ── 订单取消 ──
    ["订单取消", "order_cancel_unauthorized", "未认证取消订单", "P0",
     "/wx/order/cancel", "POST", "否", "",
     '{"orderId": 1}',
     200, 501, "", "", "", "否", ""],

    ["订单取消", "order_cancel_empty_body", "空请求体取消订单", "P1",
     "/wx/order/cancel", "POST", "是", "", "{}",
     200, 401, "", "", "", "否", ""],

    ["订单取消", "order_cancel_nonexistent", "取消不存在的订单", "P1",
     "/wx/order/cancel", "POST", "是", "",
     '{"orderId": 99999}',
     200, 402, "", "", "", "否", ""],

    # ── 确认收货 ──
    ["确认收货", "order_confirm_unauthorized", "未认证确认收货", "P0",
     "/wx/order/confirm", "POST", "否", "",
     '{"orderId": 1}',
     200, 501, "", "", "", "否", ""],

    ["确认收货", "order_confirm_empty_body", "空请求体确认收货", "P1",
     "/wx/order/confirm", "POST", "是", "", "{}",
     200, 401, "", "", "", "否", ""],

    # ── 订单删除 ──
    ["订单删除", "order_delete_unauthorized", "未认证删除订单", "P0",
     "/wx/order/delete", "POST", "否", "",
     '{"orderId": 1}',
     200, 501, "", "", "", "否", ""],

    ["订单删除", "order_delete_empty_body", "空请求体删除订单", "P1",
     "/wx/order/delete", "POST", "是", "", "{}",
     200, 401, "", "", "", "否", ""],

    ["订单删除", "order_delete_nonexistent", "删除不存在的订单", "P1",
     "/wx/order/delete", "POST", "是", "",
     '{"orderId": 99999}',
     200, 401, "", "", "", "否", ""],

    # ── 订单评价 ──
    ["订单评价", "order_comment_unauthorized", "未认证订单评价", "P0",
     "/wx/order/comment", "POST", "否", "",
     '{"orderGoodsId": 1, "content": "好评", "star": 5, "hasPicture": false, "picUrls": []}',
     200, 501, "", "", "", "否", ""],

    ["订单评价", "order_comment_empty_body", "空请求体订单评价", "P1",
     "/wx/order/comment", "POST", "是", "", "{}",
     200, 401, "", "", "", "否", ""],

    # ── 订单商品 ──
    ["订单商品", "order_goods_unauthorized", "未认证查看订单商品", "P0",
     "/wx/order/goods", "GET", "否", "ogid=1", "",
     200, 501, "", "", "", "否", ""],

    ["订单商品", "order_goods_missing_id", "不传ogid查订单商品", "P0",
     "/wx/order/goods", "GET", "是", "", "",
     200, 402, "", "ogid不能为null", "", "否", ""],

    ["订单商品", "order_goods_nonexistent", "不存在的ogid查订单商品（返回null数据）", "P1",
     "/wx/order/goods", "GET", "是", "ogid=99999", "",
     200, 0, "", "成功", "", "否", ""],

    # ── 首页 ──
    ["首页", "home_index_success", "获取首页数据（新品/热门/专题/品牌等）", "P0",
     "/wx/home/index", "GET", "否", "", "",
     200, 0, "", "成功", "", "是", "newGoodsList, hotGoodsList, banner, channel, brandList"],

    ["首页", "home_about_success", "获取关于页面信息", "P1",
     "/wx/home/about", "GET", "否", "", "",
     200, 0, "", "成功", "", "是", "name, address, phone, qq"],
]


def write_sheet(ws, cases_for_sheet: list, sheet_name: str):
    """向一个 sheet 写入表头 + 数据行"""
    # 写表头
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # 写数据
    prefix, _ = COUNTER_MAP.get(sheet_name, ("TC-UNKN", "other"))
    counter = 0

    for case_row in cases_for_sheet:
        module, case_id, desc, priority, endpoint, method, need_auth, req_params, req_body, \
            exp_status, exp_errno, exp_errno_not, exp_errmsg, exp_errmsg_contains, \
            exp_has_data, exp_check_fields = case_row

        counter += 1
        tc_number = f"{prefix}-{counter:03d}"

        row_data = [
            tc_number, case_id, module, desc, priority,
            endpoint, method, need_auth, req_params, req_body,
            exp_status, exp_errno, exp_errno_not, exp_errmsg, exp_errmsg_contains,
            exp_has_data, exp_check_fields,
        ]

        row_idx = counter + 1  # +1 for header
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = CASE_FONT
            cell.alignment = CENTER_ALIGN if col_idx in (1, 5, 7, 8, 11, 12, 13, 16) else CASE_ALIGN
            cell.border = THIN_BORDER

        # 优先级颜色
        priority_col = ws.cell(row=row_idx, column=5)
        if priority == "P0":
            priority_col.fill = P0_FILL
        elif priority == "P1":
            priority_col.fill = P1_FILL
        elif priority == "P2":
            priority_col.fill = P2_FILL

    # 列宽
    for col_idx, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 冻结首行
    ws.freeze_panes = "A2"
    # 自动筛选
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(cases_for_sheet) + 1}"


def create_excel():
    wb = openpyxl.Workbook()
    # 删除默认 sheet
    wb.remove(wb.active)

    # Sheet 顺序
    sheet_order = ["登录", "登出", "用户信息", "注册", "密码重置",
                   "商品列表", "商品详情", "商品分类", "相关商品", "商品统计",
                   "购物车列表", "购物车添加", "快速添加", "购物车更新",
                   "购物车选中", "购物车删除", "购物车数量", "购物车结算",
                   "地址列表", "地址详情", "地址保存", "地址删除",
                   "订单列表", "订单详情", "订单提交", "订单取消",
                   "确认收货", "订单删除", "订单评价", "订单商品",
                   "首页"]

    for sheet_name in sheet_order:
        ws = wb.create_sheet(sheet_name)

        # 找到属于这个 sheet 的用例
        sheet_cases = [c for c in ALL_CASES if c[0] == sheet_name]
        write_sheet(ws, sheet_cases, sheet_name)

    # ── 保存 ──
    output_dir = Path(__file__).parent.parent / "test_design"
    output_dir.mkdir(exist_ok=True)
    output_path = output_dir / "litemall_testcases.xlsx"
    wb.save(output_path)

    total = sum(1 for c in ALL_CASES)
    print(f"[OK] 已生成: {output_path}")
    print(f"   auth: {total} 例（5 个 sheet）")
    print(f"   goods: 5 个空 sheet（表头已就绪，待 Excel 中填写用例）")


if __name__ == "__main__":
    create_excel()
