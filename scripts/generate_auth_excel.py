"""
生成登录模块测试用例 Excel 文件

运行: python scripts/generate_auth_excel.py
输出: test_design/auth_module_testcases.xlsx
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path


# ============================================================
# 测试用例定义
# ============================================================
# 每个用例: (case_id, module, description, method, endpoint, need_auth,
#             request_params, request_body, expected_status, expected_errno,
#             expected_errno_not, expected_errmsg, expected_errmsg_contains,
#             expect_has_data, check_fields, priority)

CASES = {
    # ======================== 登录 ========================
    "login_success": {
        "case_id": "login_success",
        "module": "登录",
        "description": "正确的用户名和密码登录",
        "method": "POST",
        "endpoint": "/wx/auth/login",
        "need_auth": "否",
        "request_params": "",
        "request_body": '{"username": "user123", "password": "user123"}',
        "expected_status": 200,
        "expected_errno": 0,
        "expected_errno_not": "",
        "expected_errmsg": "成功",
        "expected_errmsg_contains": "",
        "expect_has_data": "是",
        "check_fields": "token, userInfo, userInfo.nickName",
        "priority": "P0",
    },
    "login_wrong_password": {
        "case_id": "login_wrong_password",
        "module": "登录",
        "description": "密码错误登录",
        "method": "POST",
        "endpoint": "/wx/auth/login",
        "need_auth": "否",
        "request_params": "",
        "request_body": '{"username": "user123", "password": "wrong_password"}',
        "expected_status": 200,
        "expected_errno": 700,
        "expected_errno_not": "",
        "expected_errmsg": "",
        "expected_errmsg_contains": "",
        "expect_has_data": "否",
        "check_fields": "",
        "priority": "P0",
    },
    "login_nonexistent_user": {
        "case_id": "login_nonexistent_user",
        "module": "登录",
        "description": "不存在的用户名登录",
        "method": "POST",
        "endpoint": "/wx/auth/login",
        "need_auth": "否",
        "request_params": "",
        "request_body": '{"username": "no_such_user_999", "password": "123456"}',
        "expected_status": 200,
        "expected_errno": 700,
        "expected_errno_not": "",
        "expected_errmsg": "",
        "expected_errmsg_contains": "账号不存在",
        "expect_has_data": "否",
        "check_fields": "",
        "priority": "P1",
    },
    "login_empty_username": {
        "case_id": "login_empty_username",
        "module": "登录",
        "description": "用户名为空登录",
        "method": "POST",
        "endpoint": "/wx/auth/login",
        "need_auth": "否",
        "request_params": "",
        "request_body": '{"username": "", "password": "user123"}',
        "expected_status": 200,
        "expected_errno": "",
        "expected_errno_not": "0",
        "expected_errmsg": "",
        "expected_errmsg_contains": "",
        "expect_has_data": "否",
        "check_fields": "",
        "priority": "P0",
    },
    "login_empty_password": {
        "case_id": "login_empty_password",
        "module": "登录",
        "description": "密码为空登录",
        "method": "POST",
        "endpoint": "/wx/auth/login",
        "need_auth": "否",
        "request_params": "",
        "request_body": '{"username": "user123", "password": ""}',
        "expected_status": 200,
        "expected_errno": "",
        "expected_errno_not": "0",
        "expected_errmsg": "",
        "expected_errmsg_contains": "",
        "expect_has_data": "否",
        "check_fields": "",
        "priority": "P0",
    },
    "login_empty_both": {
        "case_id": "login_empty_both",
        "module": "登录",
        "description": "用户名和密码都为空登录",
        "method": "POST",
        "endpoint": "/wx/auth/login",
        "need_auth": "否",
        "request_params": "",
        "request_body": '{"username": "", "password": ""}',
        "expected_status": 200,
        "expected_errno": "",
        "expected_errno_not": "0",
        "expected_errmsg": "",
        "expected_errmsg_contains": "",
        "expect_has_data": "否",
        "check_fields": "",
        "priority": "P1",
    },
    "login_username_with_spaces": {
        "case_id": "login_username_with_spaces",
        "module": "登录",
        "description": "用户名前后带空格登录（边界）",
        "method": "POST",
        "endpoint": "/wx/auth/login",
        "need_auth": "否",
        "request_params": "",
        "request_body": '{"username": "  user123  ", "password": "user123"}',
        "expected_status": 200,
        "expected_errno": "",
        "expected_errno_not": "",
        "expected_errmsg": "",
        "expected_errmsg_contains": "",
        "expect_has_data": "否",
        "check_fields": "",
        "priority": "P2",
    },
    "login_sql_injection": {
        "case_id": "login_sql_injection",
        "module": "登录",
        "description": "SQL注入尝试登录（安全测试）",
        "method": "POST",
        "endpoint": "/wx/auth/login",
        "need_auth": "否",
        "request_params": "",
        "request_body": '{"username": "admin\' OR \'1\'=\'1", "password": "admin\' OR \'1\'=\'1"}',
        "expected_status": 200,
        "expected_errno": "",
        "expected_errno_not": "0",
        "expected_errmsg": "",
        "expected_errmsg_contains": "",
        "expect_has_data": "否",
        "check_fields": "",
        "priority": "P1",
    },
    "login_xss_attempt": {
        "case_id": "login_xss_attempt",
        "module": "登录",
        "description": "XSS脚本注入尝试登录（安全测试）",
        "method": "POST",
        "endpoint": "/wx/auth/login",
        "need_auth": "否",
        "request_params": "",
        "request_body": '{"username": "<script>alert(1)</script>", "password": "<script>alert(1)</script>"}',
        "expected_status": 200,
        "expected_errno": "",
        "expected_errno_not": "0",
        "expected_errmsg": "",
        "expected_errmsg_contains": "",
        "expect_has_data": "否",
        "check_fields": "",
        "priority": "P2",
    },

    # ======================== 登出 ========================
    "logout_with_token": {
        "case_id": "logout_with_token",
        "module": "登出",
        "description": "带有效token登出",
        "method": "POST",
        "endpoint": "/wx/auth/logout",
        "need_auth": "是",
        "request_params": "",
        "request_body": "",
        "expected_status": 200,
        "expected_errno": 0,
        "expected_errno_not": "",
        "expected_errmsg": "成功",
        "expected_errmsg_contains": "",
        "expect_has_data": "否",
        "check_fields": "",
        "priority": "P0",
    },
    "logout_no_token": {
        "case_id": "logout_no_token",
        "module": "登出",
        "description": "不带token登出（应拒绝）",
        "method": "POST",
        "endpoint": "/wx/auth/logout",
        "need_auth": "否",
        "request_params": "",
        "request_body": "",
        "expected_status": 200,
        "expected_errno": 501,
        "expected_errno_not": "",
        "expected_errmsg": "请登录",
        "expected_errmsg_contains": "",
        "expect_has_data": "否",
        "check_fields": "",
        "priority": "P0",
    },
    "logout_invalid_token": {
        "case_id": "logout_invalid_token",
        "module": "登出",
        "description": "带伪造的无效token登出（应拒绝）",
        "method": "POST",
        "endpoint": "/wx/auth/logout",
        "need_auth": "否（手动设置假token）",
        "request_params": "",
        "request_body": "",
        "expected_status": 200,
        "expected_errno": 501,
        "expected_errno_not": "",
        "expected_errmsg": "请登录",
        "expected_errmsg_contains": "",
        "expect_has_data": "否",
        "check_fields": "",
        "priority": "P1",
    },

    # ======================== 用户信息 ========================
    "user_info_with_token": {
        "case_id": "user_info_with_token",
        "module": "用户信息",
        "description": "带有效token获取用户信息",
        "method": "GET",
        "endpoint": "/wx/auth/info",
        "need_auth": "是",
        "request_params": "",
        "request_body": "",
        "expected_status": 200,
        "expected_errno": 0,
        "expected_errno_not": "",
        "expected_errmsg": "成功",
        "expected_errmsg_contains": "",
        "expect_has_data": "是",
        "check_fields": "nickName, avatar, gender, mobile",
        "priority": "P0",
    },
    "user_info_no_token": {
        "case_id": "user_info_no_token",
        "module": "用户信息",
        "description": "不带token获取用户信息（应拒绝）",
        "method": "GET",
        "endpoint": "/wx/auth/info",
        "need_auth": "否",
        "request_params": "",
        "request_body": "",
        "expected_status": 200,
        "expected_errno": 501,
        "expected_errno_not": "",
        "expected_errmsg": "请登录",
        "expected_errmsg_contains": "",
        "expect_has_data": "否",
        "check_fields": "",
        "priority": "P0",
    },
    "user_info_invalid_token": {
        "case_id": "user_info_invalid_token",
        "module": "用户信息",
        "description": "带伪造的无效token获取用户信息（应拒绝）",
        "method": "GET",
        "endpoint": "/wx/auth/info",
        "need_auth": "否（手动设置假token）",
        "request_params": "",
        "request_body": "",
        "expected_status": 200,
        "expected_errno": 501,
        "expected_errno_not": "",
        "expected_errmsg": "",
        "expected_errmsg_contains": "",
        "expect_has_data": "否",
        "check_fields": "",
        "priority": "P1",
    },

    # ======================== 注册 ========================
    "register_missing_wxcode": {
        "case_id": "register_missing_wxcode",
        "module": "注册",
        "description": "缺少wxCode字段注册（wxCode需通过regCaptcha获取，当前无法自动化）",
        "method": "POST",
        "endpoint": "/wx/auth/register",
        "need_auth": "否",
        "request_params": "",
        "request_body": '{"username": "testuser_reg", "password": "Test1234!", "mobile": "13800138001"}',
        "expected_status": 200,
        "expected_errno": "",
        "expected_errno_not": "0",
        "expected_errmsg": "",
        "expected_errmsg_contains": "",
        "expect_has_data": "否",
        "check_fields": "",
        "priority": "P1",
    },
    "register_empty_username": {
        "case_id": "register_empty_username",
        "module": "注册",
        "description": "空用户名注册（应拒绝）",
        "method": "POST",
        "endpoint": "/wx/auth/register",
        "need_auth": "否",
        "request_params": "",
        "request_body": '{"username": "", "password": "Test1234!", "mobile": "13800138002"}',
        "expected_status": 200,
        "expected_errno": "",
        "expected_errno_not": "0",
        "expected_errmsg": "",
        "expected_errmsg_contains": "",
        "expect_has_data": "否",
        "check_fields": "",
        "priority": "P1",
    },
    "register_empty_password": {
        "case_id": "register_empty_password",
        "module": "注册",
        "description": "空密码注册（应拒绝）",
        "method": "POST",
        "endpoint": "/wx/auth/register",
        "need_auth": "否",
        "request_params": "",
        "request_body": '{"username": "testuser_nopwd", "password": "", "mobile": "13800138003"}',
        "expected_status": 200,
        "expected_errno": "",
        "expected_errno_not": "0",
        "expected_errmsg": "",
        "expected_errmsg_contains": "",
        "expect_has_data": "否",
        "check_fields": "",
        "priority": "P1",
    },
    "register_empty_mobile": {
        "case_id": "register_empty_mobile",
        "module": "注册",
        "description": "空手机号注册（应拒绝）",
        "method": "POST",
        "endpoint": "/wx/auth/register",
        "need_auth": "否",
        "request_params": "",
        "request_body": '{"username": "testuser_nomob", "password": "Test1234!", "mobile": ""}',
        "expected_status": 200,
        "expected_errno": "",
        "expected_errno_not": "0",
        "expected_errmsg": "",
        "expected_errmsg_contains": "",
        "expect_has_data": "否",
        "check_fields": "",
        "priority": "P1",
    },
    "register_duplicate_username": {
        "case_id": "register_duplicate_username",
        "module": "注册",
        "description": "用已存在的用户名注册（应拒绝）",
        "method": "POST",
        "endpoint": "/wx/auth/register",
        "need_auth": "否",
        "request_params": "",
        "request_body": '{"username": "user123", "password": "Test1234!", "mobile": "13800138004"}',
        "expected_status": 200,
        "expected_errno": "",
        "expected_errno_not": "0",
        "expected_errmsg": "",
        "expected_errmsg_contains": "",
        "expect_has_data": "否",
        "check_fields": "",
        "priority": "P0",
    },
    "register_invalid_mobile": {
        "case_id": "register_invalid_mobile",
        "module": "注册",
        "description": "格式非法的手机号注册（应拒绝）",
        "method": "POST",
        "endpoint": "/wx/auth/register",
        "need_auth": "否",
        "request_params": "",
        "request_body": '{"username": "testuser_badmob", "password": "Test1234!", "mobile": "12345"}',
        "expected_status": 200,
        "expected_errno": "",
        "expected_errno_not": "0",
        "expected_errmsg": "",
        "expected_errmsg_contains": "",
        "expect_has_data": "否",
        "check_fields": "",
        "priority": "P1",
    },
    "register_short_password": {
        "case_id": "register_short_password",
        "module": "注册",
        "description": "密码过短注册（应拒绝，如后端有校验）",
        "method": "POST",
        "endpoint": "/wx/auth/register",
        "need_auth": "否",
        "request_params": "",
        "request_body": '{"username": "testuser_shortpwd", "password": "12", "mobile": "13800138005"}',
        "expected_status": 200,
        "expected_errno": "",
        "expected_errno_not": "0",
        "expected_errmsg": "",
        "expected_errmsg_contains": "",
        "expect_has_data": "否",
        "check_fields": "",
        "priority": "P2",
    },

    # ======================== 密码重置 ========================
    "reset_with_valid_data": {
        "case_id": "reset_with_valid_data",
        "module": "密码重置",
        "description": "有效数据重置密码",
        "method": "POST",
        "endpoint": "/wx/auth/reset",
        "need_auth": "否",
        "request_params": "",
        "request_body": '{"password": "NewPass123!", "mobile": "13800138001", "code": "1234"}',
        "expected_status": 200,
        "expected_errno": "",
        "expected_errno_not": "",
        "expected_errmsg": "",
        "expected_errmsg_contains": "",
        "expect_has_data": "否",
        "check_fields": "",
        "priority": "P2",
    },
    "reset_empty_password": {
        "case_id": "reset_empty_password",
        "module": "密码重置",
        "description": "空密码重置（应拒绝）",
        "method": "POST",
        "endpoint": "/wx/auth/reset",
        "need_auth": "否",
        "request_params": "",
        "request_body": '{"password": "", "mobile": "13800138001", "code": "1234"}',
        "expected_status": 200,
        "expected_errno": "",
        "expected_errno_not": "0",
        "expected_errmsg": "",
        "expected_errmsg_contains": "",
        "expect_has_data": "否",
        "check_fields": "",
        "priority": "P2",
    },
    "reset_empty_mobile": {
        "case_id": "reset_empty_mobile",
        "module": "密码重置",
        "description": "空手机号重置（应拒绝）",
        "method": "POST",
        "endpoint": "/wx/auth/reset",
        "need_auth": "否",
        "request_params": "",
        "request_body": '{"password": "NewPass123!", "mobile": "", "code": "1234"}',
        "expected_status": 200,
        "expected_errno": "",
        "expected_errno_not": "0",
        "expected_errmsg": "",
        "expected_errmsg_contains": "",
        "expect_has_data": "否",
        "check_fields": "",
        "priority": "P2",
    },
}


# ============================================================
# Excel 样式定义
# ============================================================
HEADER_FONT = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

CASE_FONT = Font(name="微软雅黑", size=10)
CASE_ALIGNMENT = Alignment(horizontal="left", vertical="center", wrap_text=True)
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

P0_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")  # 浅红
P1_FILL = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")  # 浅橙
P2_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")  # 浅绿

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Sheet → 列定义
HEADERS = [
    "用例编号",
    "用例ID",
    "模块",
    "用例描述",
    "优先级",
    "接口路径",
    "请求方法",
    "需要认证",
    "请求参数",
    "请求体",
    "预期HTTP状态码",
    "预期errno",
    "预期errno不为",
    "预期errmsg",
    "预期errmsg包含",
    "预期data有值",
    "预期data字段",
]

COL_WIDTHS = [16, 28, 10, 40, 8, 30, 10, 12, 28, 55, 18, 12, 14, 22, 22, 16, 36]


def generate_excel():
    """生成 Excel 测试用例文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "认证模块测试用例"

    # ── 写表头 ──
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # ── 写数据行 ──
    module_order = ["登录", "登出", "用户信息", "注册", "密码重置"]
    sorted_cases = sorted(
        CASES.values(),
        key=lambda c: (module_order.index(c["module"]) if c["module"] in module_order else 99, c["case_id"])
    )

    counter = {"登录": 0, "登出": 0, "用户信息": 0, "注册": 0, "密码重置": 0}
    case_number_map = {
        "登录": "TC-AUTH-LOGIN",
        "登出": "TC-AUTH-LGOUT",
        "用户信息": "TC-AUTH-INFO",
        "注册": "TC-AUTH-REG",
        "密码重置": "TC-AUTH-RSET",
    }

    for row_idx, case in enumerate(sorted_cases, 2):
        mod = case["module"]
        counter[mod] += 1
        tc_number = f"{case_number_map[mod]}-{counter[mod]:03d}"

        row_data = [
            tc_number,
            case["case_id"],
            case["module"],
            case["description"],
            case["priority"],
            case["endpoint"],
            case["method"],
            case["need_auth"],
            case["request_params"],
            case["request_body"],
            case["expected_status"],
            case["expected_errno"],
            case["expected_errno_not"],
            case["expected_errmsg"],
            case["expected_errmsg_contains"],
            case["expect_has_data"],
            case["check_fields"],
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = CASE_FONT
            cell.alignment = CENTER_ALIGNMENT if col_idx in (1, 5, 7, 8, 11, 12, 13, 16) else CASE_ALIGNMENT
            cell.border = THIN_BORDER

        # 优先级颜色
        if case["priority"] == "P0":
            ws.cell(row=row_idx, column=5).fill = P0_FILL
        elif case["priority"] == "P1":
            ws.cell(row=row_idx, column=5).fill = P1_FILL
        elif case["priority"] == "P2":
            ws.cell(row=row_idx, column=5).fill = P2_FILL

    # ── 设置列宽 ──
    for col_idx, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── 冻结首行 ──
    ws.freeze_panes = "A2"

    # ── 自动筛选 ──
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(sorted_cases) + 1}"

    # ── 添加统计 Sheet ──
    ws_stats = wb.create_sheet("统计")
    stats_headers = ["模块", "P0", "P1", "P2", "合计"]
    for col_idx, h in enumerate(stats_headers, 1):
        cell = ws_stats.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    total_p0 = total_p1 = total_p2 = total_all = 0
    for row_idx, mod in enumerate(module_order, 2):
        mod_cases = [c for c in sorted_cases if c["module"] == mod]
        p0 = sum(1 for c in mod_cases if c["priority"] == "P0")
        p1 = sum(1 for c in mod_cases if c["priority"] == "P1")
        p2 = sum(1 for c in mod_cases if c["priority"] == "P2")
        total_p0 += p0
        total_p1 += p1
        total_p2 += p2
        total_all += len(mod_cases)
        for col_idx, val in enumerate([mod, p0, p1, p2, len(mod_cases)], 1):
            cell = ws_stats.cell(row=row_idx, column=col_idx, value=val)
            cell.font = CASE_FONT
            cell.alignment = CENTER_ALIGNMENT
            cell.border = THIN_BORDER

    # 合计行
    total_row = len(module_order) + 2
    for col_idx, val in enumerate(["合计", total_p0, total_p1, total_p2, total_all], 1):
        cell = ws_stats.cell(row=total_row, column=col_idx, value=val)
        cell.font = Font(name="微软雅黑", bold=True, size=10)
        cell.alignment = CENTER_ALIGNMENT
        cell.border = THIN_BORDER
        cell.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    for col_idx, width in enumerate([16, 8, 8, 8, 8], 1):
        ws_stats.column_dimensions[get_column_letter(col_idx)].width = width

    # ── 保存 ──
    output_dir = Path(__file__).parent.parent / "test_design"
    output_dir.mkdir(exist_ok=True)
    output_path = output_dir / "auth_module_testcases.xlsx"
    wb.save(output_path)
    print(f"[OK] Excel 用例文件已生成: {output_path}")
    print(f"   共 {len(sorted_cases)} 个用例")
    for mod in module_order:
        cnt = sum(1 for c in sorted_cases if c["module"] == mod)
        print(f"   - {mod}: {cnt} 个")


if __name__ == "__main__":
    generate_excel()
