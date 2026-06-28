"""
Excel 测试用例 → YAML 转换脚本

读取 test_design/litemall_testcases.xlsx（多个 sheet，每个 sheet 一个模块），
一键生成所有 YAML 测试数据文件到 tests/test_data/

用法: python scripts/excel_to_yaml.py [excel_file] [--output dir]
"""
import sys
import json
from pathlib import Path

import openpyxl
import yaml


# ============================================================
#  Sheet 名 → YAML 文件名 映射
#  格式: { sheet_name: yaml_filename_without_extension }
# ============================================================
SHEET_TO_FILE = {
    "登录":     "auth_login_data",
    "登出":     "auth_logout_data",
    "用户信息": "auth_info_data",
    "注册":     "auth_register_data",
    "密码重置": "auth_reset_data",
    "商品列表": "goods_list_data",
    "商品详情": "goods_detail_data",
    "商品分类": "goods_category_data",
    "相关商品": "goods_related_data",
    "商品统计": "goods_count_data",
    "购物车列表": "cart_index_data",
    "购物车添加": "cart_add_data",
    "快速添加": "cart_fastadd_data",
    "购物车更新": "cart_update_data",
    "购物车选中": "cart_checked_data",
    "购物车删除": "cart_delete_data",
    "购物车数量": "cart_count_data",
    "购物车结算": "cart_checkout_data",
    "地址列表": "address_list_data",
    "地址详情": "address_detail_data",
    "地址保存": "address_save_data",
    "地址删除": "address_delete_data",
    "订单列表": "order_list_data",
    "订单详情": "order_detail_data",
    "订单提交": "order_submit_data",
    "订单取消": "order_cancel_data",
    "确认收货": "order_confirm_data",
    "订单删除": "order_delete_data",
    "订单评价": "order_comment_data",
    "订单商品": "order_goods_data",
    "首页": "home_data",
}

# 模块分组（用于生成合并版 YAML）
MODULE_GROUPS = {
    "auth":  ["登录", "登出", "用户信息", "注册", "密码重置"],
    "goods": ["商品列表", "商品详情", "商品分类", "相关商品", "商品统计"],
    "cart":  ["购物车列表", "购物车添加", "快速添加", "购物车更新",
              "购物车选中", "购物车删除", "购物车数量", "购物车结算"],
    "address": ["地址列表", "地址详情", "地址保存", "地址删除"],
    "order": ["订单列表", "订单详情", "订单提交", "订单取消",
              "确认收货", "订单删除", "订单评价", "订单商品"],
    "home": ["首页"],
}

# 表头列名（固定）
HEADERS_EXPECTED = [
    "用例编号", "用例ID", "模块", "用例描述", "优先级",
    "接口路径", "请求方法", "需要认证", "请求参数", "请求体",
    "预期HTTP状态码", "预期errno", "预期errno不为",
    "预期errmsg", "预期errmsg包含", "预期data有值", "预期data字段",
]


def _parse_case(case: dict) -> dict:
    """将 Excel 一行的原始字典转为 YAML 节点"""
    expected = {}

    exp_errno = case.get("预期errno", "")
    exp_errno_not = case.get("预期errno不为", "")
    exp_errmsg = case.get("预期errmsg", "")
    exp_errmsg_contains = case.get("预期errmsg包含", "")
    exp_has_data = case.get("预期data有值", "")
    exp_check_fields = case.get("预期data字段", "")

    if exp_errno != "" and exp_errno is not None:
        val = int(exp_errno) if str(exp_errno).lstrip('-').isdigit() else exp_errno
        expected["errno"] = val
    if exp_errno_not != "" and exp_errno_not is not None:
        val = int(exp_errno_not) if str(exp_errno_not).lstrip('-').isdigit() else exp_errno_not
        expected["errno_not"] = val
    if exp_errmsg != "" and exp_errmsg is not None:
        expected["errmsg"] = str(exp_errmsg)
    if exp_errmsg_contains != "" and exp_errmsg_contains is not None:
        expected["errmsg_contains"] = str(exp_errmsg_contains)
    if exp_has_data in ("是", "true", "True"):
        expected["has_data"] = True
    elif exp_has_data in ("否", "false", "False"):
        expected["has_data"] = False
    if exp_check_fields != "" and exp_check_fields is not None:
        expected["fields"] = [f.strip() for f in str(exp_check_fields).split(",") if f.strip()]

    # request body
    req_body = case.get("请求体", "")
    request = None
    if req_body and str(req_body).strip():
        try:
            request = json.loads(str(req_body))
        except json.JSONDecodeError:
            request = {"_raw": str(req_body)}

    # URL params
    req_params = case.get("请求参数", "")
    params = None
    if req_params and str(req_params).strip():
        params = str(req_params).strip()

    node = {
        "description": str(case.get("用例描述", "")),
        "module": str(case.get("模块", "")),
        "priority": str(case.get("优先级", "")),
        "method": str(case.get("请求方法", "")).upper(),
        "need_auth": str(case.get("需要认证", "")) in ("是", "true", "True"),
        "endpoint": str(case.get("接口路径", "")),
        "expected": expected,
    }

    if params:
        node["params"] = params
    if request:
        node["request"] = request

    return node


def excel_to_yaml(excel_path: str, output_dir: str):
    """读取 Excel 的所有 sheet，按模块生成 YAML 文件 + 合并版"""
    wb = openpyxl.load_workbook(excel_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    all_module_cases = {}   # { module_name: { case_id: node } }
    total_count = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # 读取表头
        headers = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(1, col).value
            headers.append(val if val else "")

        # 读取所有数据行
        cases = []
        for row in range(2, ws.max_row + 1):
            case = {}
            for col, header in enumerate(headers, 1):
                val = ws.cell(row, col).value
                case[header] = val if val is not None else ""
            # 跳过空行
            if case.get("用例ID", ""):
                cases.append(case)

        if not cases:
            print(f"  [SKIP] {sheet_name}: 无用例数据")
            continue

        # 转换为 YAML 节点
        yaml_data = {}
        for case in cases:
            case_id = case["用例ID"]
            yaml_data[case_id] = _parse_case(case)

            # 收集到 all_module_cases 用于合并
            mod = case.get("模块", sheet_name)
            if mod not in all_module_cases:
                all_module_cases[mod] = {}
            all_module_cases[mod][case_id] = yaml_data[case_id]

        # 写分模块 YAML
        filename = SHEET_TO_FILE.get(sheet_name)
        if filename is None:
            # fallback: sheet 名拼音化
            filename = _sheet_to_filename(sheet_name)

        filepath = output_dir / f"{filename}.yaml"
        _write_yaml(filepath, yaml_data, sheet_name, excel_path)

        count = len(yaml_data)
        total_count += count
        print(f"  [OK] {sheet_name}: {count} 例 → {filepath.name}")

    # ── 生成合并版 YAML（按分组）──
    for group_name, group_sheets in MODULE_GROUPS.items():
        merged = {}
        for sheet_name in group_sheets:
            if sheet_name in all_module_cases:
                merged.update(all_module_cases[sheet_name])
        if merged:
            filepath = output_dir / f"{group_name}_data.yaml"
            _write_yaml(filepath, merged, f"{group_name} 模块合集", excel_path)
            print(f"  [OK] 合并版 {group_name}: {len(merged)} 例 → {filepath.name}")

    print(f"\n  共生成 {total_count} 个用例，{len(wb.sheetnames)} 个 sheet")


def _write_yaml(filepath: Path, data: dict, label: str, source: str):
    """写 YAML 文件"""
    with open(filepath, "w", encoding="utf-8") as f:
        f.write(f"# Litemall {label} 测试数据\n")
        f.write(f"# 由 scripts/excel_to_yaml.py 自动生成\n")
        f.write(f"# 来源: {source}\n\n")
        for idx, (case_id, node) in enumerate(data.items()):
            dump = yaml.dump(
                {case_id: node},
                default_flow_style=False,
                allow_unicode=True,
                sort_keys=False,
                indent=2,
                width=120,
            )
            f.write(dump)
            if idx < len(data) - 1:
                f.write("\n")


def _sheet_to_filename(name: str) -> str:
    """fallback: 中文 sheet 名 → 英文文件名"""
    import re
    # 简单处理：去除非字母数字，转小写
    result = re.sub(r'[^\w]', '_', name).lower()
    result = re.sub(r'_+', '_', result).strip('_')
    return result + "_data" if result else "module_data"


if __name__ == "__main__":
    args = sys.argv[1:]
    excel_file = args[0] if args else "test_design/litemall_testcases.xlsx"
    output = args[1] if len(args) > 1 else "tests/test_data"

    excel_path = Path(excel_file)
    if not excel_path.is_absolute():
        excel_path = Path(__file__).parent.parent / excel_file

    if not excel_path.exists():
        print(f"[ERROR] Excel 文件不存在: {excel_path}")
        sys.exit(1)

    print(f"Excel: {excel_path}")
    print(f"Output: {output}\n")
    excel_to_yaml(str(excel_path), output)
